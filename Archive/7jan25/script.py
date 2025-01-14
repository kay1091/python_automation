import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
import os

# --- Configuration ---
SFID_FILE = "input/SFID_file.xlsx"
SFDC_DUMP_FILE = "input/SFDC_dump.xlsx"
TEMPLATE_FILE = "input/Weekly_Template.xlsx"
OUTPUT_FILE = "output/Updated_Template.xlsx"
TEMPLATE_SHEET_NAME = "SFDC"
LARGE_DEAL_THRESHOLD = 20000000
FISCAL_OFFSET = 3  # Month offset for fiscal year

# --- Helper Functions ---
def calculate_opportunity_status_from_template(proposal_status, stage, created_date_str):
    if proposal_status in ["No-Go", "On-Hold", "Deferred"]:
        return "No-BID"
    if stage == "7 - Contract Award":
        return "WON"
    if stage == "Lost":
        return "LOST"
    if stage in ["1 - Opportunity", "2 - Qualification", "3 - Pursuit", "4 - Proposal", "5 - Closing", "6 - Verbal"]:
        return "OPEN"
    try:
        created_date = pd.to_datetime(created_date_str)
        if pd.isnull(created_date) or (pd.Timestamp.today() - created_date).days > 90:
            return "CLOSED"
    except (TypeError, ValueError):
        return "OPEN" #Default to open if date parsing fails


def calculate_large_deal_from_value(amount):
    if pd.isna(amount):
        return "--"
    if amount >= LARGE_DEAL_THRESHOLD:
        return "Yes"
    elif amount > 0:
        return "No"
    return "--"

def get_last_monday(date_value):
    if pd.isnull(date_value):
        return None
    date_obj = pd.to_datetime(date_value).date()
    start_of_week = date_obj - timedelta(days=date_obj.weekday())
    return start_of_week

def calculate_fiscal_year_short(date_value, fiscal_offset=FISCAL_OFFSET):
    if pd.isnull(date_value):
        return None
    date = pd.to_datetime(date_value)
    year = date.year
    month = date.month
    if month > fiscal_offset:
        return f"FY{year % 100}"
    else:
        return f"FY{(year - 1) % 100}"

def calculate_quarter(date_value):
    if pd.isnull(date_value):
        return None
    month = pd.to_datetime(date_value).month
    if 4 <= month <= 6:
        return "Q1"
    elif 7 <= month <= 9:
        return "Q2"
    elif 10 <= month <= 12:
        return "Q3"
    elif 1 <= month <= 3:
        return "Q4"
    return None

def calculate_week_from_date(date_value):
    if pd.isnull(date_value):
        return None
    return pd.to_datetime(date_value).isocalendar()[1]

def calculate_bid_director(group_sbu):
    """
    Determine the Bid Director based on the Group SBU value.
    """
    if group_sbu in ["GM APAC", "GM ASIA", "GM ANZ"]:
        return "Piyush J"
    elif group_sbu in ["GM EME", "GM MIDDLE EAST", "GM CONTINENTAL EUROPE", "GM UNITED KINGDOM"]:
        return "Samrat B"
    elif group_sbu in ["BET NA TELCO", "BET NA EMERGING", "BET NA BFS US", "BET NA CANADA", "Platinum ac-Citi", "Platinum ac-JPMC"]:
        return "Nadeem A"
    elif group_sbu in ["HIL LIFE SCIENCES", "HIL HEALTHCARE", "HIL INSURANCE"]:
        return "Anish R"
    elif group_sbu in ["TIME ALPHABET", "TIME IME", "TIME TECHNOLOGY"]:
        return "Vineeth V"
    else:
        return "-"

# --- Step 1: Load Input Files ---
try:
    sfid_df = pd.read_excel(SFID_FILE)
    sfdc_dump_df = pd.read_excel(SFDC_DUMP_FILE)
except FileNotFoundError as e:
    print(f"Error: Could not find input files. Please ensure they are in the 'input' directory. Error: {e}")
    exit()

# Normalize Columns - added error handling for inconsistent column names
try:
    sfid_df.columns = [str(col).strip() for col in sfid_df.columns]
    sfdc_dump_df.columns = [str(col).strip() for col in sfdc_dump_df.columns]
except KeyError as e:
    print(f"Error: Inconsistent column names in input files. Please check your Excel files. Error: {e}")
    exit()


# --- Step 2: Combine Data ---
# No merging, instead iterate over the two dataframes based on index
template_data = []
template_dict = {
    "Account Name": "",
    "SFID": "",
    "Opportunity ID": "",
    "Created Date": None,
    "Opportunity Name": "",
    "Opportunity Description": "",
    "Group SBU": None,
    "Created By": None,
    "Stage": None,
    "Est Deal Value in USD": None,
    "Opp Type": None,
    "Vertical Practice": None,
    "Tech. Practice": None,
    "Service Offering": None,
    "Engagement Type": None,
    "Probability": None,
    "Close Date": None,
    "Next Steps": None,
    "Loss Stage": None,
    "Lost Reason": None,
    "Age": None,
    "BOLT": None,
    "Doc. Recvd. Date": None,
    "Category": None,
    "Partner Details": None,
    "Proposed Sub. Date": None,
    "Domain Practice": None,
    "Tech Practice": None,
    "Solution SPOCs": None,
    "Delivery SPOC": None,
    "Proposal Owner": None,
    "Allocation% Proposal Owner 1": None,
    "Proposal Owner 2": None,
    "Allocation% Proposal Owner 2": None,
    "Proposal Writer": None,
    "Allocation% Proposal Writer 1": None,
    "Proposal Writer 2": None,
    "Allocation% Proposal Writer 2": None,
    "Orals SPOC": None,
    "Bid Director": None,
    "Proposal Updates": None,
    "Proposal Status": None,
    "Actual Sub. Date": None,
    "Commercial Value": None,
    "DSC": None,
    "Opportunity Stage": None,
    "Post Sub. Activity": None,
    "PSA Activity Status": None,
    "PSA Activity Cls. Date": None,
    "PSA Activity Update": None,
    "Est. Deal Value": None,
    "Large Deal": None,
    "SBU Mapping": None,
    "Created in Week": None,
    "Submitted in Week": None,
    "Closing in Week": None,
    "PSA Comp. in Week": None,
    "Sub. Dt. Mapping": None,
    "Cl. Dt. Mapping": None,
    "PSA Comp. Dt. Mapping": None,
    "Opp. Status": None,
    "Sb. FY": None,
    "Sb. Qtr.": None,
    "Cl. FY": None,
    "Cl. QTR": None,
    "TCV Brk. Up": None,
    "BFS/ETS": None,
    "Direct/ Related": None,
    "OB Related Status": None,
    "TCV Related Status": None,
}


for i in range(max(len(sfid_df), len(sfdc_dump_df))):
    template_row = {}

    # Get data from both dataframes based on index
    sfid_row = sfid_df.iloc[i] if i < len(sfid_df) else {}
    sfdc_row = sfdc_dump_df.iloc[i] if i < len(sfdc_dump_df) else {}

    for col, default in template_dict.items():
         if col == "Account Name":
            if "Account Name" in sfid_row and pd.notna(sfid_row.get("Account Name")):
                 value = sfid_row["Account Name"]
            elif "Account Name" in sfdc_row and pd.notna(sfdc_row.get("Account Name")):
                 value = sfdc_row["Account Name"]
            else:
                value = default
         elif col == "SFID":
              if "SFID" in sfid_row and pd.notna(sfid_row.get("SFID")):
                  value = sfid_row["SFID"]
              elif "Opportunity ID" in sfdc_row and pd.notna(sfdc_row.get("Opportunity ID")):
                 value = sfdc_row["Opportunity ID"]
              else:
                   value = default
         elif col == "Created Date":
             if "Created Date" in sfid_row and pd.notna(sfid_row.get("Created Date")):
                  value = sfid_row["Created Date"]
             elif "Created Date" in sfdc_row and pd.notna(sfdc_row.get("Created Date")):
                   value = sfdc_row["Created Date"]
             else:
                  value = default

         elif col == "Opportunity Name" and "Opportunity Name" in sfid_row:
             value = sfid_row.get("Opportunity Name")

         elif col == "Opportunity Description":
              if "Opportunity Description" in sfid_row and pd.notna(sfid_row.get("Opportunity Description")):
                   value = sfid_row["Opportunity Description"]
              elif "Description" in sfdc_row and pd.notna(sfdc_row.get("Description")):
                   value = sfdc_row["Description"]
              else:
                    value = default

         elif col == "Group SBU":
            if "Group SBU" in sfdc_row and pd.notna(sfdc_row.get("Group SBU")):
                value = sfdc_row["Group SBU"]
            else:
                 value = default
         elif col == "Created By":
            if "Opportunity Owner" in sfdc_row and pd.notna(sfdc_row.get("Opportunity Owner")):
                 value = sfdc_row["Opportunity Owner"]
            else:
                 value = default
         elif col == "Stage":
              if "Deal Stage" in sfid_row and pd.notna(sfid_row.get("Deal Stage")):
                   value = sfid_row["Deal Stage"]
              elif "Stage" in sfdc_row and pd.notna(sfdc_row.get("Stage")):
                   value = sfdc_row["Stage"]
              else:
                    value = default

         elif col == "Est Deal Value in USD":
           if "Amount (converted)" in sfdc_row and pd.notna(sfdc_row.get("Amount (converted)")):
                 value = sfdc_row["Amount (converted)"]
           elif "$ Value (M)" in sfid_row and pd.notna(sfid_row.get("$ Value (M)")):
                 value = sfid_row["$ Value (M)"]
           else:
                 value = default

         elif col == "Opp Type":
               if "Type" in sfdc_row and pd.notna(sfdc_row.get("Type")):
                   value = sfdc_row["Type"]
               else:
                    value = default
         elif col == "Vertical Practice":
              if "Vertical Practice" in sfdc_row and pd.notna(sfdc_row.get("Vertical Practice")):
                   value = sfdc_row["Vertical Practice"]
              else:
                    value = default
         elif col == "Tech. Practice":
              if "Partner Details" in sfid_row and pd.notna(sfid_row.get("Partner Details")):
                 value = sfid_row["Partner Details"]
              else:
                 value = default
         elif col == "Service Offering":
              if "Service Offering" in sfdc_row and pd.notna(sfdc_row.get("Service Offering")):
                   value = sfdc_row["Service Offering"]
              else:
                   value = default
         elif col == "Engagement Type":
              if "Project Type" in sfdc_row and pd.notna(sfdc_row.get("Project Type")):
                   value = sfdc_row["Project Type"]
              else:
                    value = default

         elif col == "Probability":
             if "Probability (%)" in sfdc_row and pd.notna(sfdc_row.get("Probability (%)")):
                   value = sfdc_row["Probability (%)"]
             else:
                    value = default

         elif col == "Close Date":
            if "Close Date" in sfid_row and pd.notna(sfid_row.get("Close Date")):
               value = sfid_row["Close Date"]
            elif "Close Date" in sfdc_row and pd.notna(sfdc_row.get("Close Date")):
                 value = sfdc_row["Close Date"]
            else:
                value = default

         elif col == "Next Steps":
             if "Next Step" in sfdc_row and pd.notna(sfdc_row.get("Next Step")):
                 value = sfdc_row["Next Step"]
             else:
                   value = default
         elif col == "Loss Stage":
            if "Loss Stage" in sfdc_row and pd.notna(sfdc_row.get("Loss Stage")) :
               value = sfdc_row["Loss Stage"]
            else:
                  value = default
         elif col == "Lost Reason":
               if "Lost Reason" in sfdc_row and pd.notna(sfdc_row.get("Lost Reason")):
                     value = sfdc_row["Lost Reason"]
               else:
                     value = default
         elif col == "Age":
            if "Age" in sfdc_row and pd.notna(sfdc_row.get("Age")):
               value = sfdc_row["Age"]
            else:
                 value = default
         elif col == "BOLT":
              if "BOLT Details" in sfdc_row and pd.notna(sfdc_row.get("BOLT Details")):
                    value = sfdc_row["BOLT Details"]
              else:
                    value = default
         elif col == "Category":
              if "Activity Type" in sfid_row and pd.notna(sfid_row.get("Activity Type")):
                    value = sfid_row["Activity Type"]
              else:
                  value = default
         elif col == "Partner Details":
                if "Partner Details" in sfid_row and pd.notna(sfid_row.get("Partner Details")):
                     value = sfid_row["Partner Details"]
                else:
                     value = default
         elif col == "Proposed Sub. Date":
              if "Due Date" in sfid_row and pd.notna(sfid_row.get("Due Date")):
                     value = sfid_row["Due Date"]
              else:
                  value = default
         elif col == "Solution SPOCs":
             if "Solution SPOCs" in sfid_row and pd.notna(sfid_row.get("Solution SPOCs")):
                   value = sfid_row["Solution SPOCs"]
             else:
                  value = default

         elif col == "Delivery SPOC":
              if "Delivery Lead" in sfid_row and pd.notna(sfid_row.get("Delivery Lead")):
                    value = sfid_row["Delivery Lead"]
              else:
                    value = default

         elif col == "Proposal Owner":
             if "Bid Manager" in sfid_row and pd.notna(sfid_row.get("Bid Manager")):
                   value = sfid_row["Bid Manager"]
             else:
                   value = default
         elif col == "Proposal Writer":
                if "Proposal Writer" in sfid_row and pd.notna(sfid_row.get("Proposal Writer")):
                      value = sfid_row["Proposal Writer"]
                else:
                     value = default

         elif col == "Orals SPOC":
              if "Orals SPOC" in sfid_row and pd.notna(sfid_row.get("Orals SPOC")):
                    value = sfid_row["Orals SPOC"]
              else:
                    value = default

         elif col == "Bid Director":
            value = calculate_bid_director(template_row.get("Group SBU"))
         elif col == "Proposal Updates":
                if "Status/ Next Steps" in sfid_row and pd.notna(sfid_row.get("Status/ Next Steps")):
                     value = sfid_row["Status/ Next Steps"]
                else:
                    value = default
         elif col == "Proposal Status":
            if "Deal Status" in sfid_row and pd.notna(sfid_row.get("Deal Status")):
                  value = sfid_row["Deal Status"]
            else:
                value = default

         elif col == "Actual Sub. Date":
                if "Due Date" in sfid_row and pd.notna(sfid_row.get("Due Date")):
                     value = sfid_row["Due Date"]
                else:
                     value = default
         elif col == "Commercial Value":
           if "Amount (converted)" in sfdc_row and pd.notna(sfdc_row.get("Amount (converted)")):
                 value = sfdc_row["Amount (converted)"]
           elif "$ Value (M)" in sfid_row and pd.notna(sfid_row.get("$ Value (M)")):
                 value = sfid_row["$ Value (M)"]
           else:
                 value = default
         elif col == "DSC":
              if "DSC Status" in sfid_row and pd.notna(sfid_row.get("DSC Status")):
                    value = sfid_row["DSC Status"]
              else:
                    value = default
         elif col == "Opportunity Stage":
             if "Stage" in sfdc_row and pd.notna(sfdc_row.get("Stage")):
                   value = sfdc_row["Stage"]
             else:
                   value = default
         elif col == "Est. Deal Value":
            if "Amount (converted)" in sfdc_row and pd.notna(sfdc_row.get("Amount (converted)")):
                 value = sfdc_row["Amount (converted)"]
            else:
                 value = None

         elif col == "SBU Mapping":
              if "Group SBU" in sfdc_row and pd.notna(sfdc_row.get("Group SBU")):
                  value = sfdc_row["Group SBU"]
              else:
                    value = default
         else:
                value = default #Use default value if it is not in either of the dataframes
         template_row[col] = value

     #Specific calculations and logic
    template_row["Doc. Recvd. Date"] = get_last_monday(datetime.today())

    #Convert to float if present
    if template_row.get("Est Deal Value") is not None:
       try:
           template_row["Est Deal Value"] = float(template_row["Est Deal Value"])
       except (ValueError, TypeError):
          template_row["Est Deal Value"] = None


    template_row["Est Deal Value in USD"] = template_row.get("Est Deal Value")
    template_row["Commercial Value"] = template_row.get("Est Deal Value")
    template_row["Large Deal"] = calculate_large_deal_from_value(template_row.get("Est Deal Value"))
    template_row["Created in Week"] = calculate_week_from_date(template_row.get("Created Date"))
    template_row["Submitted in Week"] = calculate_week_from_date(template_row.get("Proposed Sub. Date"))
    template_row["Closing in Week"] = calculate_week_from_date(template_row.get("Close Date"))
    template_row["Opp. Status"] = calculate_opportunity_status_from_template(template_row.get("Proposal Status"), template_row.get("Stage"), template_row.get("Created Date"))
    template_row["Sb. FY"] = calculate_fiscal_year_short(template_row.get("Proposed Sub. Date"))
    template_row["Sb. Qtr."] = calculate_quarter(template_row.get("Proposed Sub. Date"))
    template_row["Cl. FY"] = calculate_fiscal_year_short(template_row.get("Close Date"))
    template_row["Cl. QTR"] = calculate_quarter(template_row.get("Close Date"))
    template_data.append(template_row)

# --- Step 4: Load and Update Template ---
try:
    os.makedirs("output", exist_ok=True) # Create output directory if it doesn't exist
    template_wb = openpyxl.load_workbook(TEMPLATE_FILE)
    template_ws = template_wb[TEMPLATE_SHEET_NAME]

    # Get existing data
    existing_data = []
    for row in template_ws.iter_rows(min_row=2, values_only=True):
         existing_data.append(list(row))

    template_ws.delete_rows(2, template_ws.max_row)  # Clear existing data
    #Append existing data back
    for row in existing_data:
         template_ws.append(row)

    for row in dataframe_to_rows(pd.DataFrame(template_data), index=False, header=False):
        template_ws.append(row)

    template_wb.save(OUTPUT_FILE)
    print(f"Template updated successfully: {OUTPUT_FILE}")
except FileNotFoundError:
    print(f"Error: Could not find template file '{TEMPLATE_FILE}'. Please ensure it exists in the 'input' directory.")
except Exception as e:
    print(f"An error occurred: {e}")
    exit()