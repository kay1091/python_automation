import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers
import argparse
from datetime import datetime

def load_and_normalize_data(file_path, id_column, file_type):
    """Loads and normalizes data from an Excel file.

    Args:
        file_path (str): Path to the Excel file.
        id_column (str): Name of the ID column to normalize.
        file_type (str): Type of file ("SFID" or "SFDC") for logging.

    Returns:
        pandas.DataFrame: The loaded and normalized DataFrame.
    """
    try:
        df = pd.read_excel(file_path)
        if df.empty:
            raise ValueError(f"{file_type} file is empty.")
        if id_column in df.columns:
            df[id_column] = df[id_column].astype(str).str.strip().str.upper()
        print(f"Loaded and normalized {file_type} file.")
        return df
    except FileNotFoundError:
        raise FileNotFoundError(f"{file_type} file not found at {file_path}")
    except Exception as e:
        raise Exception(f"Error reading {file_type} file: {e}")

def merge_dataframes(sfid_df, sfdc_dump_df):
    """Merges two dataframes.

    Args:
        sfid_df (pandas.DataFrame): SFID DataFrame.
        sfdc_dump_df (pandas.DataFrame): SFDC Dump DataFrame.

    Returns:
        pandas.DataFrame: Merged DataFrame.
    """
    merged_df = pd.merge(sfid_df, sfdc_dump_df, left_on="SFID", right_on="Opportunity ID", how="left")
    print("Merged SFID and SFDC Dump data.")
    return merged_df

def load_and_clear_template(template_file, sheet_name):
    """Loads and clears the specified sheet in the Excel template.

    Args:
        template_file (str): Path to the template file.
        sheet_name (str): Name of the worksheet.

    Returns:
        openpyxl.worksheet.worksheet.Worksheet: Template worksheet object.
    """
    try:
        template_wb = openpyxl.load_workbook(template_file)
        if sheet_name not in template_wb.sheetnames:
            raise ValueError(f"Worksheet '{sheet_name}' not found in template file.")
        template_ws = template_wb[sheet_name]
        if template_ws.max_row > 1:
           template_ws.delete_rows(2, template_ws.max_row)
        print(f"Loaded template workbook and cleared data in sheet '{sheet_name}'.")
        return template_ws, template_wb
    except FileNotFoundError:
        raise FileNotFoundError(f"Template file not found at {template_file}")
    except Exception as e:
        raise Exception(f"Error loading template file: {e}")

def write_data_to_template(merged_df, template_ws, template_column_mapping):
    """Writes data from the merged dataframe to the template worksheet.

    Args:
        merged_df (pandas.DataFrame): Merged DataFrame.
        template_ws (openpyxl.worksheet.worksheet.Worksheet): Template worksheet object.
        template_column_mapping (dict): Dictionary for mapping the column names.
    """

    # Get column names from merged dataframe and convert to lowercase for comparison
    merged_cols_lower = [col.lower() for col in merged_df.columns]

    # Prepare rows
    for index, row in merged_df.iterrows():
        row_values = []
        for template_col, source_col in template_column_mapping.items():
            if source_col and source_col.lower() in merged_cols_lower:
                value = row[source_col]
                if isinstance(value, datetime):
                    row_values.append(value.strftime("%Y-%m-%d")) # Format date before adding
                else:
                   row_values.append(value)

            else:
                row_values.append(None)  # Handle missing data

        template_ws.append(row_values)

    # Set date format for the specified columns
    for col_idx, template_col in enumerate(template_column_mapping.keys()):
         if template_col in ["Created Date", "Proposed Sub. Date", "Actual Sub. Date"]:
              for cell in template_ws.iter_cols(min_col=col_idx+1, max_col=col_idx+1, min_row=2):
                  for c in cell:
                      c.number_format = "yyyy-mm-dd"

def save_updated_template(template_wb, output_file):
    """Saves the updated template to a file.

    Args:
        template_wb (openpyxl.workbook.workbook.Workbook): Template workbook object.
        output_file (str): Path to the output file.
    """
    try:
        template_wb.save(output_file)
        print(f"Updated template saved successfully to {output_file}!")
    except Exception as e:
        raise Exception(f"Error saving updated template: {e}")


def main():
    """Main function to execute the data processing."""

    parser = argparse.ArgumentParser(description="Merge data from Excel files into a template.")
    parser.add_argument("--sfid_file", default="input/SFID_file.xlsx", help="Path to SFID file.")
    parser.add_argument("--sfdc_file", default="input/SFDC_dump.xlsx", help="Path to SFDC dump file.")
    parser.add_argument("--template_file", default="input/Weekly_Template.xlsx", help="Path to template file.")
    parser.add_argument("--output_file", default="output/Updated_Template.xlsx", help="Path to output file.")
    parser.add_argument("--template_sheet", default="SFDC Data", help="Name of the template sheet.")
    args = parser.parse_args()

    # --- Configuration ---
    SFID_FILE = args.sfid_file
    SFDC_DUMP_FILE = args.sfdc_file
    TEMPLATE_FILE = args.template_file
    OUTPUT_FILE = args.output_file
    TEMPLATE_SHEET_NAME = args.template_sheet

    # --- Step 1: Load and Normalize Data (SFID_File) ---
    try:
        sfid_df = load_and_normalize_data(SFID_FILE, "SFID", "SFID")
    except Exception as e:
        print(f"Error with SFID file: {e}")
        exit()

    # --- Step 2: Load and Normalize Data (SFDC_Dump) ---
    try:
        sfdc_dump_df = load_and_normalize_data(SFDC_DUMP_FILE, "Opportunity ID", "SFDC Dump")
    except Exception as e:
        print(f"Error with SFDC Dump file: {e}")
        exit()

    # --- Step 3: Merge DataFrames ---
    merged_df = merge_dataframes(sfid_df, sfdc_dump_df)

    # --- Step 4: Load Template Workbook ---
    try:
         template_ws, template_wb = load_and_clear_template(TEMPLATE_FILE, TEMPLATE_SHEET_NAME)
    except Exception as e:
        print(f"Error with Template file: {e}")
        exit()
    
    # --- Step 5: Define Column Mapping ---
    template_column_mapping = {
        "Account Name": "Account Name_x",
        "SFID": "SFID",
        "Created Date": "Created Date_x",
        "Opportunity Name": "Opportunity Name_x",
        "Opportunity Description": "Opportunity Description",
        "Group SBU": "SBU",
        "Created By": "Created By_x",
        "Proposed Sub. Date": "Due Date",
        "Domain Practice": "Practice",
        "Tech Practice": "Technology/Skills",
        "Solution SPOCs": "Solution SPOCs",
        "Delivery SPOC": "Delivery Lead",
        "Proposal Owner": "Bid Manager",
        "Proposal Writer": "Proposal Writer",
        "Orals SPOC": "Orals SPOC",
        "Bid Director": "Client Partner",
        "Proposal Updates": "Status/Next Steps",
        "Proposal Status": "Deal Status",
        "Actual Sub. Date": "Close Date",
        "Commercial Value": "$ Value (M)",
        "DSC": "DSC Status",
        "Opportunity Stage": "Deal Stage",
        "Est. Deal Value": "Amount",
        "Large Deal": None,
        "SBU Mapping": None,
        "Opp. Status": None,
        "Sb. FY": None,
        "Sb. Qtr.": None,
        "Cl. FY": "Fiscal Period",
        "Cl. QTR": None,
        "TCV Brk. Up": None,
        "Direct/ Related": None,
        "OB Related Status": None,
        "TCV Related Status": None,
        }

    # --- Step 6: Write Data to Template ---
    write_data_to_template(merged_df, template_ws, template_column_mapping)

    # --- Step 7: Save Updated Template ---
    try:
        save_updated_template(template_wb, OUTPUT_FILE)
    except Exception as e:
        print(f"Error with output file: {e}")

if __name__ == "__main__":
    main()